#!/usr/bin/env python3
"""
Import ride plans from the Team Asha Excel spreadsheet into Supabase.

Usage:
    DATABASE_URL='postgresql://...' python scripts/import_ride_plans.py /path/to/spreadsheet.xlsx
"""

import os
import re
import sys
import psycopg2
import psycopg2.extras
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("pip install openpyxl")
    sys.exit(1)

DATABASE_URL = os.environ.get('DATABASE_URL')
if not DATABASE_URL:
    # Try loading from .env
    env_path = Path(__file__).parent.parent / '.env'
    if env_path.exists():
        for line in env_path.read_text().splitlines():
            if line.startswith('DATABASE_URL='):
                DATABASE_URL = line.split('=', 1)[1].strip()
                break

if not DATABASE_URL:
    print("DATABASE_URL not set")
    sys.exit(1)

# Sheets to skip (not actual ride plans)
SKIP_SHEETS = {
    'asha pbp strava details',
    'WNY 1000K - plan',  # Personal plan, not a team ride plan
}


def slugify(name):
    """Convert sheet name to a URL-friendly slug."""
    s = name.lower().strip()
    s = re.sub(r'[^a-z0-9]+', '-', s)
    s = s.strip('-')
    return s


def safe_float(v):
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def safe_int(v):
    f = safe_float(v)
    return int(f) if f is not None else None


def detect_stop_type(location):
    loc = location.lower()
    if 'start' in loc and 'finish' not in loc:
        return 'start'
    elif 'finish' in loc:
        return 'finish'
    elif 'control' in loc:
        return 'control'
    elif any(w in loc for w in ['water', 'refill', 'snack', 'lunch', 'dinner',
                                 'food', 'break', 'coffee', 'epp selfie']):
        return 'rest'
    else:
        return 'waypoint'


def find_plan_sheets(wb):
    """Identify sheets that contain ride plan data."""
    plan_sheets = []
    for name in wb.sheetnames:
        if name in SKIP_SHEETS:
            continue
        ws = wb[name]
        row1 = [str(c.value or '').strip().lower() for c in ws[1][:11]]
        has_distance = 'distance' in row1
        has_elevation = any('elevation' in v for v in row1)
        has_segment = any('segment' in v for v in row1)

        if has_distance and (has_elevation or has_segment):
            plan_sheets.append(name)
    return plan_sheets


def parse_plan_sheet(ws):
    """Parse a ride plan worksheet into stops and metadata."""
    stops = []
    rwgps_urls = []

    for r in range(2, min(ws.max_row + 1, 60)):
        loc = ws.cell(r, 1).value
        dist = ws.cell(r, 2).value
        elev = ws.cell(r, 3).value
        seg_time = ws.cell(r, 6).value
        notes_col = 11 if ws.max_column >= 11 else ws.max_column
        notes = ws.cell(r, notes_col).value

        if not loc or not str(loc).strip():
            continue

        loc_str = str(loc).strip()

        # Capture RWGPS URLs
        if loc_str.startswith('http'):
            label = str(ws.cell(r, 2).value or '').strip()
            rwgps_urls.append({'url': loc_str, 'label': label})
            continue

        # Only include rows with a valid distance
        dist_f = safe_float(dist)
        if dist_f is None:
            continue

        stops.append({
            'location': loc_str[:200],
            'stop_type': detect_stop_type(loc_str),
            'distance_miles': round(dist_f, 1),
            'elevation_gain': safe_int(elev),
            'segment_time_min': safe_int(seg_time),
            'notes': str(notes)[:500] if notes else None,
        })

    if len(stops) < 3:
        return None, None, None

    total_dist = max((s['distance_miles'] for s in stops if s['distance_miles']), default=0)
    total_elev = sum(s['elevation_gain'] or 0 for s in stops)

    # Identify RWGPS URLs
    rwgps_official = None
    rwgps_team = None
    for u in rwgps_urls:
        label = u['label'].lower()
        if 'team' in label or 'asha' in label or 'control' in label:
            rwgps_team = u['url']
        elif not rwgps_official:
            rwgps_official = u['url']

    return stops, {
        'total_distance_miles': round(total_dist, 1),
        'total_elevation_ft': total_elev,
        'rwgps_url': rwgps_official,
        'rwgps_url_team': rwgps_team,
    }, rwgps_urls


def main():
    if len(sys.argv) < 2:
        print("Usage: python scripts/import_ride_plans.py <spreadsheet.xlsx>")
        sys.exit(1)

    xlsx_path = sys.argv[1]
    if not os.path.exists(xlsx_path):
        print(f"File not found: {xlsx_path}")
        sys.exit(1)

    print(f"Loading {xlsx_path}...")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    plan_sheet_names = find_plan_sheets(wb)
    print(f"Found {len(plan_sheet_names)} ride plan sheets\n")

    conn = psycopg2.connect(DATABASE_URL)
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    imported = 0
    skipped = 0

    for sheet_name in plan_sheet_names:
        ws = wb[sheet_name]
        stops, meta, _ = parse_plan_sheet(ws)

        if not stops:
            print(f"  SKIP {sheet_name} (too few stops)")
            skipped += 1
            continue

        slug = slugify(sheet_name)

        # Upsert ride_plan
        cur.execute("""
            INSERT INTO ride_plan (name, slug, total_distance_miles, total_elevation_ft, rwgps_url, rwgps_url_team)
            VALUES (%s, %s, %s, %s, %s, %s)
            ON CONFLICT (slug) DO UPDATE SET
                name = EXCLUDED.name,
                total_distance_miles = EXCLUDED.total_distance_miles,
                total_elevation_ft = EXCLUDED.total_elevation_ft,
                rwgps_url = EXCLUDED.rwgps_url,
                rwgps_url_team = EXCLUDED.rwgps_url_team
            RETURNING id
        """, (sheet_name, slug, meta['total_distance_miles'], meta['total_elevation_ft'],
              meta['rwgps_url'], meta['rwgps_url_team']))
        plan_id = cur.fetchone()['id']

        # Delete old stops and re-insert
        cur.execute("DELETE FROM ride_plan_stop WHERE ride_plan_id = %s", (plan_id,))

        for i, stop in enumerate(stops):
            cur.execute("""
                INSERT INTO ride_plan_stop (ride_plan_id, stop_order, location, stop_type,
                    distance_miles, elevation_gain, segment_time_min, notes)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """, (plan_id, i + 1, stop['location'], stop['stop_type'],
                  stop['distance_miles'], stop['elevation_gain'],
                  stop['segment_time_min'], stop['notes']))

        print(f"  OK   {sheet_name} -> /{slug} ({len(stops)} stops, {meta['total_distance_miles']}mi, {meta['total_elevation_ft']:,}ft)")
        imported += 1

    conn.commit()
    conn.close()

    print(f"\nDone: {imported} imported, {skipped} skipped")


if __name__ == '__main__':
    main()
